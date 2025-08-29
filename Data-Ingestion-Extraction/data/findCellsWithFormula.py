import openpyxl

# Load the Excel file
workbook = openpyxl.load_workbook('Clean-CombinedNewTemplate.xlsx')

# Loop through all the sheets in the workbook
for sheet_name in workbook.sheetnames:
    sheet = workbook[sheet_name]

    # Loop through all cells in the sheet
    for row in sheet.iter_rows():
        for cell in row:
            if cell.data_type == 'f':
                print(f"Sheet: {sheet_name}, Cell {cell.coordinate} contains a formula: {cell.formula}")

# Close the Excel file when you're done
workbook.close()
