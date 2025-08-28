import openpyxl

# Define the workbook name and the list of expected column names
workbook_name = 'CombinedFile-CSAMTemplate.xlsx'
expected_columns = ['Identifier or Host Name', 'IP Address (Internal)',	'IP Address (External)',	'NAT IPs',	'AD Domain',	'CPU Core',	'Memory (GB)',	'Drive Space (GB)',	'OS Name',	'OS Version',	'Lifecycle',	'Location',	'Hosting/CSP Contract',	'Asset Category',	'Asset Type',	'Virtual',	'Hardware Make',	'Hardware Model',	'Manufacturer Serial Number',	'BIOS UUID/GUID',	'MAC Address(es)',	'Public',	'High Value Asset',	'GFE',	'Date Device Added to System Boundary',	'System Owner / Device Manager',	'Device Operator',	'Primary System Boundary CSAM Acronym',	'Primary System Boundary CSAM ID',	'Systems Supported CSAM Acronym',	'System Supported CSAM ID',	'First Tier Supplier']


def check_and_insert_columns_in_worksheet(worksheet, expected_columns):
    """
    Check if the first row of the worksheet contains all the expected columns.
    If not, insert the missing columns at the end of the row.
    """
    # Get the first row values
    first_row = [cell.value for cell in worksheet[1]]

    # Find missing columns
    missing_columns = [col for col in expected_columns if col not in first_row]

    # Insert missing columns at the end of the first row
    if missing_columns:
        # Find the last column index
        last_column_index = len(first_row) + 1
        for idx, col in enumerate(missing_columns, start=last_column_index):
            worksheet.cell(row=1, column=idx, value=col)
        return missing_columns
    return None


def main():
    # Load the workbook
    wb = openpyxl.load_workbook(workbook_name)

    # Iterate through all the worksheets in the workbook
    for sheet_name in wb.sheetnames:
        ws = wb[sheet_name]
        missing_columns = check_and_insert_columns_in_worksheet(ws, expected_columns)

        if missing_columns:
            print(f"Worksheet '{sheet_name}' was missing columns: {missing_columns}. They have been added.")
        else:
            print(f"Worksheet '{sheet_name}' contains all the expected columns.")

    # Save the workbook after making changes
    wb.save('CorrectedCombinedFile-CSAMTemplate.xlsx')


if __name__ == "__main__":
    main()
