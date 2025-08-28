import openpyxl

# Function to check if a row contains a specific string
def row_contains_string(row, target_strings):
    for cell in row:
        if cell.value and any(target in str(cell.value) for target in target_strings):
            return True
    return False

# Load the workbook
workbook = openpyxl.load_workbook('Raw-CombinedFile-CSAMTemplate.xlsx')

# Iterate through each sheet in the workbook
for sheet_name in workbook.sheetnames:
    sheet = workbook[sheet_name]

    # List of row numbers to delete (1-based index)
    rows_to_delete = []

    # Strings to check for in the cells
    target_strings = ["Operating System Name", "bpvhxwviis271", "Hardware Inventory", "Critical"]

    # Iterate through the rows and mark for deletion if they contain the target strings
    for row_num, row in enumerate(sheet.iter_rows(), start=1):
        if row_contains_string(row, target_strings):
            rows_to_delete.append(row_num)
        if row_num >= 10:
            break

    # Sort the rows to delete in descending order
    rows_to_delete.sort(reverse=True)

    # Delete the rows
    for row_num in rows_to_delete:
        sheet.delete_rows(row_num)
    print(f"Sheet '{sheet_name}' Completed")

# Save the modified workbook
workbook.save('CombinedFile-CSAMTemplate.xlsx')

# Check the 1st row of each workbook.
# Define the values to check for in any cell of the first row
values_to_check = ["IP Address (Internal)", "NAT IPs", "Identifier or Host Name", "CPU Core"]

# Open the Excel workbook
workbook2 = openpyxl.load_workbook('CombinedFile-CSAMTemplate.xlsx')

# Iterate through each sheet in the workbook
for sheet_name in workbook2.sheetnames:
    sheet = workbook2[sheet_name]

    found_values = []

    # Iterate through the cells in the first row of the sheet
    for cell in sheet[1]:
        if cell.value in values_to_check:
            found_values.append(cell.value)

    if found_values:
        print(f"Sheet '{sheet_name}' Found")
    else:
        print(f"Sheet '{sheet_name}' does not have any of the expected values in the first row.")

# Close the workbook
workbook2.close()
