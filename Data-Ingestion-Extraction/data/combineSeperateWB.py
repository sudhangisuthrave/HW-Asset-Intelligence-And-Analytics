import openpyxl

# Define the source workbook and the target workbook
source_workbook_path = 'CombinedFileAll.xlsx'
target_workbook_path = 'CombinedNewTemplate.xlsx'

# Load the source workbook
source_workbook = openpyxl.load_workbook(source_workbook_path)

# Create a target workbook if it doesn't exist
try:
    target_workbook = openpyxl.load_workbook(target_workbook_path)
except FileNotFoundError:
    target_workbook = openpyxl.Workbook()

# Specify the string to search for
search_string = 'BIOS UUID'

# Loop through sheets in the source workbook
for sheet_name in source_workbook.sheetnames:
    source_sheet = source_workbook[sheet_name]

    # Flag to check if the search string is found in the sheet
    string_found = False

    # Loop through all cells in the sheet
    for row in source_sheet.iter_rows():
        for cell in row:
            if search_string in str(cell.value):
                string_found = True
                break  # Exit the inner loop if the string is found
        if string_found:
            break  # Exit the outer loop if the string is found

    # If the search string is found, create a copy of the source sheet in the target workbook
    if string_found:
        target_sheet = target_workbook.create_sheet(sheet_name)

        # Copy data from source to target sheet
        for row in source_sheet.iter_rows():
            for cell in row:
                if cell.value is not None:
                    target_sheet[cell.coordinate] = cell.value
    print(f"Sheet: {sheet_name} Completed")

# Save the target workbook
target_workbook.save(target_workbook_path)

# Close both workbooks
source_workbook.close()
target_workbook.close()
