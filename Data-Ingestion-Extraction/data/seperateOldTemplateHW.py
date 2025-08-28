import openpyxl

# Define the source workbook and the target workbooks
source_workbook_path = 'CombinedFileNewTemplate.xlsx'
with_string_workbook_path = 'NewTemplateFile.xlsx'
without_string_workbook_path = 'OldTemplateFile.xlsx'

# Load the source workbook
source_workbook = openpyxl.load_workbook(source_workbook_path)

# Create target workbooks if they don't exist
try:
    with_string_workbook = openpyxl.load_workbook(with_string_workbook_path)
except FileNotFoundError:
    with_string_workbook = openpyxl.Workbook()

try:
    without_string_workbook = openpyxl.load_workbook(without_string_workbook_path)
except FileNotFoundError:
    without_string_workbook = openpyxl.Workbook()

# Specify the string to search for
search_string = 'BIOS UUID'

# Loop through sheets in the source workbook
for sheet_name in source_workbook.sheetnames:
    source_sheet = source_workbook[sheet_name]

    # Flag to check if the search string is found in the sheet
    string_found = False

    # Loop through all cells in the sheet
    for row in source_sheet.iter_rows():
        for cell in row:
            if search_string in str(cell.value):
                string_found = True
                break  # Exit the inner loop if the string is found
        if string_found:
            break  # Exit the outer loop if the string is found

    # Create a copy of the source sheet in the appropriate target workbook
    if string_found:
        target_workbook = with_string_workbook
    else:
        target_workbook = without_string_workbook

    target_sheet = target_workbook.create_sheet(sheet_name)

    # Copy data from source to target sheet
    for row in source_sheet.iter_rows():
        for cell in row:
            target_sheet[cell.coordinate] = cell.value

# Save the target workbooks
with_string_workbook.save(with_string_workbook_path)
without_string_workbook.save(without_string_workbook_path)

# Close all workbooks
source_workbook.close()
with_string_workbook.close()
without_string_workbook.close()
