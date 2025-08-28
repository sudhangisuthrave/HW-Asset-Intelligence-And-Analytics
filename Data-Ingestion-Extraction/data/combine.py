import time
import os
from openpyxl import Workbook

from openpyxl import load_workbook

dir_containing_files = "C:\\Users\\Sudhangi.Suthrave\\PycharmProjects\\inventory\\csam_inventory\\data"
dest_wb = Workbook()

for root, dir, filenames in os.walk(dir_containing_files):
    for file in filenames:
        if file.endswith('.xlsx'):
            start = time.perf_counter()
            file_name = file.split('.')[0]
            '''The absolute path for all HW inventory files'''
            file_path = os.path.abspath(os.path.join(root, file))

            '''Create a new sheet each time you have to append a xlsx file'''
            dest_wb.create_sheet(file_name)
            dest_ws = dest_wb[file_name]

            '''Read all of the source data'''
            source_wb = load_workbook(file_path)
            '''Include all worksheets from a workbook'''
            for source_sheet in source_wb.worksheets:
                for row in source_sheet.rows:
                    for cell in row:
                        '''Make sure to not include empty rows. Empty rows increase processing time exponentially'''
                        if cell.value is not None:
                            dest_ws[cell.coordinate] = cell.value

                '''Calculate elapsed time for each xlsx file in seconds'''
                elapsed_time = time.perf_counter() - start
                print(f"File: {file_name} Time Elapsed: {elapsed_time:0.4f} secs")

'''Save all the data in a final CombinedFile.xlsx workbook'''
dest_wb.save("CombinedFileSWAM.xlsx")
