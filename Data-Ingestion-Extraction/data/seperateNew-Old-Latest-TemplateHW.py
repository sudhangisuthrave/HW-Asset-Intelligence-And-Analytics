import openpyxl

# Create blank target workbooks
with_string_workbook = openpyxl.Workbook()
csam_string_workbook = openpyxl.Workbook()
without_string_workbook = openpyxl.Workbook()


# Define the source workbook and the target workbooks
source_workbook_path = 'CombinedAllTemplateFiles.xlsx'
with_string_workbook_path = 'Raw-CombinedFile-NewTemplate.xlsx'
csam_string_workbook_path = 'Raw-CombinedFile-CSAMTemplate.xlsx'
without_string_workbook_path = 'Raw-CombinedFile-OldTemplate.xlsx'

# Load the source workbook
source_workbook = openpyxl.load_workbook(source_workbook_path)

# Create target workbooks if they don't exist
try:
    with_string_workbook = openpyxl.load_workbook(with_string_workbook_path)
except FileNotFoundError:
    with_string_workbook = openpyxl.Workbook()

try:
    csam_string_workbook = openpyxl.load_workbook(csam_string_workbook_path)
except FileNotFoundError:
    csam_string_workbook = openpyxl.Workbook()

try:
    without_string_workbook = openpyxl.load_workbook(without_string_workbook_path)
except FileNotFoundError:
    without_string_workbook = openpyxl.Workbook()

# Specify the strings to search for
search_strings = ['BIOS UUID', 'Systems Supported CSAM Acronym']

# Loop through sheets in the source workbook
for sheet_name in source_workbook.sheetnames:
    source_sheet = source_workbook[sheet_name]

    # Flags to check if each search string is found in the sheet
    strings_found = [False] * len(search_strings)

    # Loop through all cells in the sheet
    for row in source_sheet.iter_rows():
        for cell in row:
            for i, search_string in enumerate(search_strings):
                if search_string in str(cell.value):
                    strings_found[i] = True

    # Create a copy of the source sheet in the appropriate target workbook(s)
    if all(strings_found):  # If both 'BIOS UUID' and 'Systems Supported CSAM Acronym' are found
        target_workbook = csam_string_workbook
    elif strings_found[0]:  # If only 'BIOS UUID' is found
        target_workbook = with_string_workbook
    else:
        target_workbook = without_string_workbook

    target_sheet = target_workbook.create_sheet(sheet_name)

    # Copy data from source to target sheet
    for row in source_sheet.iter_rows():
        for cell in row:
            target_sheet[cell.coordinate] = cell.value

# Save the target workbooks
print("Saving all 3 seperate workbooks.")
with_string_workbook.save(with_string_workbook_path)
csam_string_workbook.save(csam_string_workbook_path)
without_string_workbook.save(without_string_workbook_path)

# Close all workbooks
source_workbook.close()
with_string_workbook.close()
csam_string_workbook.close()
without_string_workbook.close()
