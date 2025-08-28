import openpyxl

# Function to check if a row contains a specific string
def row_contains_string(row, target_strings):
    for cell in row:
        if cell.value and any(target in str(cell.value) for target in target_strings):
            return True
    return False

# Load the workbook
workbook = openpyxl.load_workbook('CombinedNewTemplate.xlsx')

# Iterate through each sheet in the workbook
for sheet_name in workbook.sheetnames:
    sheet = workbook[sheet_name]

    # List of row numbers to delete (1-based index)
    rows_to_delete = [1, 3, 4, 5]

    # Strings to check for in the cells
    target_strings = ["EDUPTCVPP005", "FL-EDSW-01"]

    # Iterate through the rows and mark for deletion if they contain the target strings
    for row_num, row in enumerate(sheet.iter_rows(), start=1):
        if row_contains_string(row, target_strings):
            rows_to_delete.append(row_num)

    # Sort the rows to delete in descending order
    rows_to_delete.sort(reverse=True)

    # Delete the rows
    for row_num in rows_to_delete:
        sheet.delete_rows(row_num)

# Save the modified workbook
workbook.save('Clean-CombinedNewTemplate.xlsx')
