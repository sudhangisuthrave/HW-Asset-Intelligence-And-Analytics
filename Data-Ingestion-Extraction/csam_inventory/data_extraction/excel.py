"""Extract data from excel files"""
from collections.abc import Sequence
from pathlib import Path
from typing import Dict, Generator, List, Union
from zipfile import BadZipfile

from openpyxl import load_workbook
import pandas as pd

from csam_inventory.data_extraction.utils import clean_hostname
from csam_inventory.log import LoggingBase


# potential hostname header field names, from most likely to least likely
HEADERS = [
    'hostname',
    'identifier',
    'host name',
    'inventory',
    'doed name',
    'fims hostname',
    'doed vm name',
    'server name',
    'node',
    'asset classification 1',
    'configuration item'
]

# certain sheets have hostnames in columns other than a column with
# 'hostname' in the header but these headers shouldn't necessarily
# be added to the HEADERS list since they may appear in other
# workbooks and NOT contain hostname data
UNIQUE_HEADERS = {
    'hw-inventory-349': 'IP Address',
    'hw-inventory-527': 'Name/Type',
    'hw-inventory-534': 'Master Host'
}

# ignore these headers in all files
EXCLUDE_HEADERS = [
    'prior hostname',
    'server name / function'
]

# ignore these headers for specific files
# useful if a file has multiple matching headers
EXCLUDE_HEADERS_BY_FILE = {
    'hw-inventory-593': ['doed vm name']
}

# ignore these sheets
EXCLUDE_SHEETS = [
    'decommissioned'
]

# ignore rows where the first or second cells contain the following
EXCLUDE_ROW_START = [
    "assets are owned and maintained by",
    "delete column a",
    "guidance",
    "if a device has multiple ip addresses",
    "infrastructure example"
    "mandatory or optional",
    "optional",
    "valid values",
    "yes or no",
    'reviewed',
    'customer',
    'contract name',
    'siebel',
    'award date',
    'start date',
    'contract end date',
    'inventory system',
    'principle office',
    'csam name',
    'server order',
    'invoice name',
    'server order form'
]


class ExcelProcessor(LoggingBase):
    """Extract hostnames from Excel files"""

    def _load_workbook(self, file_path: str) -> Dict[str, pd.DataFrame]:
        """Load an Excel workbook for processing

        Parameters
        ----------
        file_path: str
            path to Excel file to load

        Returns
        -------
        Dict[str, pd.DataFrame]
            a dictionary where each key represents a sheet from the workbook and
            the corresponding values are pandas DataFrames that contain the
            sheet data
        """
        self.logger.info("Loading workbook at %s.", file_path)
        workbook_data = {}  # type: Dict[str, pd.DataFrame]

        try:
            # need to load with data_only=True otherwise cells with formulas
            # will return formulas rather than computed values
            workbook = load_workbook(file_path, data_only=True)

        except BadZipfile:
            # xlsx a docx file are compressed zip files, if they are password
            # protected, a BadZipfile exception is raised
            self.logger.warning(
                "Unable to open workbook, possibly password protected."
            )

            return workbook_data

        for sheet_name in workbook.sheetnames:
            if sheet_name.lower() in EXCLUDE_SHEETS:
                continue

            sheet = workbook[sheet_name]
            data = sheet.values

            # find header row
            headers_found = False
            columns = []
            while not headers_found:
                try:
                    row = next(data)

                # if iteration through rows stops, we've reached the end of the
                # sheet and should continue to the next one
                except StopIteration:
                    break

                if not ExcelProcessor._is_good_row(row, is_header=True):
                    continue

                columns = row
                headers_found = True

                # print(f"---{sheet_name}---")
                # print(columns)
                # print(f"---{sheet_name}---")

            if not columns:
                self.logger.warning("Could not find header row in sheet %s.",
                                    sheet_name)

                continue

            # process remaining rows
            good_rows = []

            for row in data:
                # ignore rows with empty values
                # or rows that start with certain values
                has_entries = any(str(x).strip() for x in row if x)

                correct_length = len(columns)

                is_good_row = ExcelProcessor._is_good_row(
                    row,
                    correct_length,
                    is_header=False
                )

                if has_entries and is_good_row:
                    good_rows.append(row)

            if not good_rows:
                self.logger.warning("Could not find data in sheet %s.",
                                    sheet_name)

                continue

            workbook_data[sheet_name] = pd.DataFrame(good_rows, columns=columns)

        return workbook_data

    @staticmethod
    def _is_good_row(row: Sequence, length: int = 2,
                     is_header: bool = True) -> bool:
        """Check if a row of data is suitable as a possible header row or data
        row

        Parameters
        ----------
        row: Generator
            row of data

        length: int
            desired length of entries; when checking header rows, the length of
            the header row must be greater than or equal to this value; when
            checking data rows, the length of the row must be equal to this
            value

        is_header: bool
            determines which length check to use

        Returns
        -------
        bool
            True if row meets testing conditions, False otherwise
        """
        # ignore initial rows that start with certain values
        row_len = len(row)
        entry_0 = str(row[0]).lower().strip() if row_len > 0 and row[0] else ""
        entry_1 = str(row[1]).lower().strip() if row_len > 1 and row[1] else ""

        has_bad_entries = (
                any(entry_0.startswith(x) for x in EXCLUDE_ROW_START)
                or any(entry_1.startswith(x) for x in EXCLUDE_ROW_START)
        )

        if is_header:
            has_correct_length = (len(set(str(x).lower().strip()
                                          for x in row if x)) >= length)
        else:
            has_correct_length = row_len == length

        if not has_bad_entries and has_correct_length:
            return True

        return False

    def _extract_hosts(self, file_path: str,
                       workbook_data: Dict[str, pd.DataFrame]) -> List[str]:
        """Extract hostnames from workbook data

        Parameters
        ----------
        file_path: str
            path to the Excel file

        workbook_data: Dict[str, pd.DataFrame]
            a dictionary where each key represents a sheet from the workbook and
            the corresponding values are pandas DataFrames that contain the
            sheet data

        Returns
        -------
        List[str]
            a list of hostnames, possibly empty
        """
        hostnames = []
        file_path = str(file_path)  # in case it's passed in as a Path object
        file_stem = Path(file_path).stem

        for sheet_name, system_data in workbook_data.items():
            self.logger.info("Processing sheet %s.", sheet_name)

            if file_stem in UNIQUE_HEADERS.keys():
                hostname_column = UNIQUE_HEADERS[file_stem]

            else:
                hostname_column = self._find_hostname_column(
                    file_path,
                    system_data
                )

            if not hostname_column:
                self.logger.warning("Could not find hostname column in %s. "
                                    "Manual modification may be required.",
                                    sheet_name)

                continue

            hosts = [x for x in system_data[hostname_column] if x]

            # print(f"---{sheet_name}---")
            # print(hosts)
            # print(f"---{sheet_name}---")

            for value in hosts:
                hostnames.extend(clean_hostname(value))

        hostnames = list(set(hostnames))
        self.logger.info('Found %d hosts in %s.', len(hostnames), file_path)
        return hostnames

    def _find_hostname_column(self, file_path: str,
                              data_frame: pd.DataFrame) -> Union[str, None]:
        """Locate the column in a DataFrame with a header in the HEADERS list

        Parameters
        ----------
        data_frame: pandas DataFrame
            DataFrame representing a sheet from a workbook

        Returns
        -------
        str
            the name of the matching column or None
        """
        file_stem = Path(file_path).stem
        excluded_headers = EXCLUDE_HEADERS_BY_FILE.get(file_stem, [])

        for name in data_frame.columns:
            if not name or name.lower().strip() in excluded_headers:
                continue

            name_in_headers = any(x in name.lower() for x in HEADERS)
            name_in_excluded = any(x in name.lower() for x in EXCLUDE_HEADERS)

            if name_in_headers and not name_in_excluded:
                self.logger.debug("Found header: %s.", name)
                return name

        return None

    def process_inventory(self, workbook_path: str) -> List[str]:
        """Process an Excel file to extract hostname data

        Parameters
        ----------
        workbook_path: str
            path to Excel file

        Returns
        -------
        List[str]
            a list of hostnames, possibly empty
        """
        inventory_data = self._load_workbook(workbook_path)
        hostnames = self._extract_hosts(workbook_path, inventory_data)
        return hostnames


def extract_hostnames(file_path: str) -> List[str]:
    """Extract hostnames from an Excel file

    This is a helper method that handles creation of an instance of the
    ExcelProcessor class.

    Parameters
    ----------
    file_path: str
        path to an Excel file

    Returns
    -------
    List[str]
        a list of hostnames, possibly empty
    """
    processor = ExcelProcessor()
    return processor.process_inventory(file_path)
